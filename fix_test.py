import re

with open("tests/test_real_formats.py", "r") as f:
    content = f.read()

replacement = """
    # We use openpyxl to build a multi-sheet with metadata
    from openpyxl import Workbook
    wb = Workbook()
    
    # Sheet 1: Formatted (should be ignored)
    ws1 = wb.active
    ws1.title = 'Formatted Report'
    ws1['A1'] = 'This is a formatted report with no usable data'
    
    # Sheet 2: Raw Data
    ws2 = wb.create_sheet('Raw Data Report')
    ws2['B2'] = 'Abhishek Pal - FML Ad account Report'
    ws2['C2'] = 'This is raw data. See Sheet 1 for a formatted report.'
    ws2['E2'] = 'Report Period: Aug 1, 2026 - Aug 7, 2026'
    
    headers = [
        "Campaign name", "Ad set name", "Ad name", "Delivery status", "Delivery level", 
        "Attribution setting", "Result type", "Results", "Reach", "Frequency", 
        "Cost per result", "Amount spent (INR)", "Quality ranking", "Engagement rate ranking", 
        "Conversion rate ranking", "Impressions", "CPM (cost per 1,000 impressions)", 
        "Link clicks", "Shop clicks", "CPC (cost per link click)", "CTR (link click-through rate)", 
        "Clicks (all)", "CTR (all)", "CPC (all)", "Landing page views", "Cost per landing page view", 
        "Reporting starts", "Reporting ends"
    ]
    
    # Write headers on row 3 (openpyxl is 1-indexed, so row=3), starting at column B (column=2)
    for col_num, data in enumerate(headers, start=2):
        ws2.cell(row=3, column=col_num, value=data)
        
    # Write data row
    data_row = [
        "FML-C2-WN", "IT INTERESTS BROAD", "Yacht > There are literally Millions IT jobs Available - 002 - Copy 4", 
        "active", "ad", "7-day click", "Website leads", 10, 1000, 1.0, 100, 1000.0, 
        "-", "-", "-", 1000, 100, 10, 0, 10, 1.0, 10, 1.0, 10, 10, 10, "2026-01-01", "2026-01-31"
    ]
    for col_num, data in enumerate(data_row, start=2):
        ws2.cell(row=4, column=col_num, value=data)
        
    wb.save(str(path))
"""

pattern = re.compile(r'    # We use xlsxwriter.*?workbook\.close\(\)', re.DOTALL)
content = pattern.sub(replacement, content)

with open("tests/test_real_formats.py", "w") as f:
    f.write(content)
