import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, NumberFormatDescriptor
from typing import Dict

def apply_formatting(ws):
    # Bold headers and freeze top row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    
    # Auto-fit columns (simple approximation)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50

def generate_workbook(filepath: str, data: Dict[str, pd.DataFrame]):
    """
    Generate the 12-sheet Excel workbook.
    data keys should match the exact sheet names.
    """
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    sheet_names = [
        "1. ⚙ Settings & Run Log",
        "2. 📋 All Sales (Attributed)",
        "3. 📢 Campaign Summary",
        "4. 🎯 Ad Set Summary",
        "5. 🎨 Ad Creative Summary",
        "6. 🏦 Ad Account Comparison",
        "7. 💰 Free vs Paid Funnel",
        "8. 🚨 Zero-ROI Waste Report",
        "9. ❓ Unattributed Sales",
        "10. 🔁 Old Leads (Separate)",
        "11. 💳 Spend Reference",
        "12. ✅ Verification",
        "13. 🚫 Excluded Sales"
    ]
    
    # Currency and Percentage formats
    currency_fmt = '₹#,##0.00'
    pct_fmt = '0.00%'
    roas_fmt = '0.0"x"'
    
    for sheet_name in sheet_names:
        ws = wb.create_sheet(title=sheet_name[:31]) # Excel limits to 31 chars
        
        df = data.get(sheet_name, pd.DataFrame())
        
        # Replace NaN/Inf with None for Excel
        df = df.replace([float('inf'), float('-inf')], None)
        df = df.where(pd.notnull(df), None)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            
        apply_formatting(ws)
        
        # Apply specific number formats if columns can be identified
        # This is a simple heuristic based on column headers
        if r_idx > 1:
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                name_lower = str(col_name).lower()
                
                fmt = None
                if 'revenue' in name_lower or 'spend' in name_lower or 'profit' in name_lower or 'amount' in name_lower or 'cpl' in name_lower or 'cac' in name_lower or 'fee' in name_lower:
                    fmt = currency_fmt
                elif 'roas' in name_lower:
                    fmt = roas_fmt
                elif '%' in name_lower or 'rate' in name_lower:
                    fmt = pct_fmt
                    
                if fmt:
                    for row_idx in range(2, r_idx + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = fmt
                            
                        # Conditional formatting for Profit/ROAS/Waste
                        if 'profit' in name_lower and isinstance(cell.value, (int, float)):
                            if cell.value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                            elif cell.value < 0:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
                        
                        if 'status' in name_lower and isinstance(cell.value, str):
                            if cell.value == 'FAIL':
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            elif cell.value == 'PASS':
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif cell.value == 'WARNING':
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    wb.save(filepath)
    return filepath
