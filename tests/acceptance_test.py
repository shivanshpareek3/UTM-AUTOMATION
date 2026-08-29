import pandas as pd
import openpyxl
import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.ingestion import read_file
from src.pipeline import run_pipeline

def run_acceptance():
    print("--- STARTING ACCEPTANCE TEST ---")
    leads_file = '/Users/apple/Downloads/20260825_071521_GlobalJobMasterclass1530328_subscriber.csv'
    sales_file = '/Users/apple/Downloads/22 and 23 Aug sales - Copy of sale (1).csv'
    meta1_file = '/Users/apple/Downloads/FML-X-ABHISHEK-PAL-Campaigns-15-Aug-2026-21-Aug-2026.csv'
    meta2_file = '/Users/apple/Downloads/A-hishek-Pal---FML-Campaigns-15-Aug-2026-21-Aug-2026.csv'

    leads_df = read_file(leads_file)
    sales_df = read_file(sales_file).dropna(how='all')
    meta_df1 = read_file(meta1_file)
    meta_df2 = read_file(meta2_file)

    print("1. Manual mapping only — zero auto-mapping (Simulated UI mapping dict)")
    leads_map = {
        'email': 'email',
        'registration_date': 'created_at',
        'campaign': 'utm_campaign',
        'ad_set': 'utm_medium',
        'ad_creative': 'utm_source',
        'name': 'first_name',
        'phone': 'phone'
    }
    
    sales_map = {
        'email': 'email',
        'name': 'name',
        'phone': 'phone'
    }
    
    meta_map1 = {
        'campaign': 'Campaign name',
        'spend': 'Amount spent (INR)',
        'Date': 'Reporting starts'
    }
    
    meta_map2 = {
        'campaign': 'Campaign name',
        'spend': 'Amount spent (INR)',
        'Date': 'Reporting starts'
    }

    inv_leads = {v: k for k, v in leads_map.items()}
    inv_sales = {v: k for k, v in sales_map.items()}
    inv_meta1 = {v: ('Day' if k == 'Date' else k) for k, v in meta_map1.items()}
    inv_meta2 = {v: ('Day' if k == 'Date' else k) for k, v in meta_map2.items()}

    leads_df.rename(columns=inv_leads, inplace=True)
    sales_df.rename(columns=inv_sales, inplace=True)
    meta_df1.rename(columns=inv_meta1, inplace=True)
    meta_df2.rename(columns=inv_meta2, inplace=True)

    print("3. Leads mapping includes name:", 'name' in leads_df.columns)
    print("4. Sales mapping includes name and phone:", 'name' in sales_df.columns and 'phone' in sales_df.columns)
    print("5. Meta mapping uses Date:", 'Day' in meta_df1.columns)

    settings = {
        'start_date': '2026-08-15',
        'cutoff_date': '2026-08-21',
        'lead_sales_start_date': '2026-08-15',
        'lead_sales_end_date': '2026-08-21',
        'meta_start_date': '2026-08-15',
        'meta_end_date': '2026-08-21',
        'sale_date_source': 'Actual Sale Date', # simulated fallback since no sale date
        'amount_source': 'Fallback Price Per Sale', # simulated fallback since no amount
        'payment_status_source': 'Treat All Imported Sales as Successful',
        'paid_markers': ["paid", "cpc", "cpm", "ppc", "paid_social", "paid_search", "google", "facebook", "instagram", "meta", "linkedin", "youtube", "bing", "snapchat", "twitter", "ads", "advertisement"],
        'client_name': 'Acceptance Test',
        'report_name': 'Acceptance Test',
        'fallback_price': 8999.0,
        'zero_roi_threshold': 5000.0,
        'funnel_type': 'Paid',
        'paid_funnel_price': 8999.0
    }

    output_path = 'output/acceptance_report.xlsx'
    os.makedirs('output', exist_ok=True)
    
    try:
        metrics, ver_df, xl_path = run_pipeline(leads_df, sales_df, [meta_df1, meta_df2], settings, output_path)
        print("8. Generate Report must never crash with KeyError:", True)
    except Exception as e:
        print("8. Generate Report crashed with exception:", e)
        return

    print("\n--- FINAL METRICS ---")
    print(f"Total Leads: {metrics.get('total_leads')}")
    print(f"Total Sales: {metrics.get('total_sales')}")
    print(f"Attributed Sales: {metrics.get('attributed_sales')}")
    print(f"Revenue: {metrics.get('total_revenue')}")
    print(f"Registration Revenue: {metrics.get('total_reg_revenue')}")
    print(f"Raw Meta Spend: {metrics.get('raw_meta_spend')}")
    print(f"Attributed Spend: {metrics.get('attributed_spend')}")
    print(f"Unallocated Spend: {metrics.get('unallocated_spend')}")
    print(f"Profit: {metrics.get('profit')}")
    print(f"ROAS: {metrics.get('roas')}")
    print(f"ROI: {metrics.get('roi_percent')}")
    print(f"CAC: {metrics.get('cac')}")
    print(f"CPL: {metrics.get('cpl')}")
    
    print("\n--- EXCEL WORKBOOK CHECKS ---")
    wb = openpyxl.load_workbook(xl_path)
    sheet_names = wb.sheetnames
    
    print("24. Final workbook contains Campaign Summary:", any("Campaign Summary" in s for s in sheet_names))
    
    unwanted = ["Unattributed Sales", "Old Leads", "Old Sales", "Spend Reference", "Verification", "Excluded Sales", "Free vs Paid Funnel", "Zero-ROI Waste Report", "Ad Account Comparison"]
    found_unwanted = [s for s in sheet_names if any(u in s for u in unwanted)]
    print(f"25. Final workbook unwanted sheets found: {found_unwanted}")
    
    camp_sheet = [s for s in sheet_names if "Campaign Summary" in s]
    if camp_sheet:
        ws = wb[camp_sheet[0]]
        headers = [cell.value for cell in ws[1]]
        print(f"23. Excel column headers for Campaign Summary: {headers}")
    else:
        print("23. Campaign Summary sheet not found.")

if __name__ == "__main__":
    run_acceptance()
