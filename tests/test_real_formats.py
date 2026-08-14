import pytest
import pandas as pd
import os
from src.pipeline import run_pipeline
from src.ingestion import read_file
from src.inspection import load_aliases

def create_real_leads(tmp_path):
    leads = pd.DataFrame({
        "first_name": ["Prabir"],
        "email": ["m.prabirda@gmail.com"],
        "phone": [918851764064],
        "city": ["Gurgaon"],
        "what's_your_current_it": ["Manager "],
        "utm_source": ["FML-C2-WN"],
        "utm_medium": ["IT INTERESTS BROAD"],
        "utm_campaign": ["Facebook_Mobile_Reels"],
        "utm_term": [120237677610220521],
        "utm_content": ["Yacht > There are literally Millions IT jobs Available - 002 - Copy 4"],
        "utm_id": [120237677609990521],
        "touch_point": [3],
        "utm_flow": ["FML-C2-WN"],
        "ffu_id": [8299],
        "gclid": [None],
        "fbclid": ["IwZXh0"],
        "source_url": ["--"],
        "ip": ["223.233.68.45"],
        "created_at": ["2026-01-01 00:31:00"]
    })
    path = tmp_path / "Lead Sheet Abhishek pal .csv"
    leads.to_csv(path, index=False)
    return path

def create_real_sales(tmp_path):
    sales = pd.DataFrame({
        "name": ["priti"],
        "email": ["m.prabirda@gmail.com"],  # Match lead for test
        "phone": [8951119168]
        # Notice: missing sale_date, order_amount, payment_status
    })
    path = tmp_path / "Sales .csv"
    sales.to_csv(path, index=False)
    return path

def create_real_meta(tmp_path):
    # Meta format 2: Sheet 1 is Formatted, Sheet 2 is Raw Data with shifted headers
    path = tmp_path / "Abhishek-Pal-FML-Ad-account-Report.xlsx"
    

    # We use openpyxl to build a multi-sheet with metadata
    from openpyxl import Workbook
    wb = Workbook()
    
    # Sheet 1: Formatted (should be ignored)
    ws1 = wb.active
    ws1.title = 'Formatted Report'
    ws1['A1'] = 'This is a formatted report with no usable data'
    
    # Sheet 2: Raw Data
    ws2 = wb.create_sheet('Raw Data Report')
    ws2['B2'] = 'Abhishek Pal - FML Ad account Report'
    ws2['C2'] = 'This is raw data. See Sheet 1 for a formatted report.'
    ws2['E2'] = 'Report Period: Aug 1, 2026 - Aug 7, 2026'
    
    headers = [
        "Campaign name", "Ad set name", "Ad name", "Delivery status", "Delivery level", 
        "Attribution setting", "Result type", "Results", "Reach", "Frequency", 
        "Cost per result", "Amount spent (INR)", "Quality ranking", "Engagement rate ranking", 
        "Conversion rate ranking", "Impressions", "CPM (cost per 1,000 impressions)", 
        "Link clicks", "Shop clicks", "CPC (cost per link click)", "CTR (link click-through rate)", 
        "Clicks (all)", "CTR (all)", "CPC (all)", "Landing page views", "Cost per landing page view", 
        "Reporting starts", "Reporting ends"
    ]
    
    # Write headers on row 3 (openpyxl is 1-indexed, so row=3), starting at column B (column=2)
    for col_num, data in enumerate(headers, start=2):
        ws2.cell(row=3, column=col_num, value=data)
        
    # Write data row
    data_row = [
        "FML-C2-WN", "IT INTERESTS BROAD", "Yacht > There are literally Millions IT jobs Available - 002 - Copy 4", 
        "active", "ad", "7-day click", "Website leads", 10, 1000, 1.0, 100, 1000.0, 
        "-", "-", "-", 1000, 100, 10, 0, 10, 1.0, 10, 1.0, 10, 10, 10, "2026-01-01", "2026-01-31"
    ]
    for col_num, data in enumerate(data_row, start=2):
        ws2.cell(row=4, column=col_num, value=data)
        
    wb.save(str(path))

    return path

def test_real_formats_ingestion_and_mapping(tmp_path):
    leads_path = create_real_leads(tmp_path)
    sales_path = create_real_sales(tmp_path)
    meta_path = create_real_meta(tmp_path)
    
    leads_df = read_file(str(leads_path))
    sales_df = read_file(str(sales_path))
    meta_df = read_file(str(meta_path))
    
    # Assert smart ingestion found the meta headers on row 2 (which is header=2)
    assert 'Campaign name' in meta_df.columns
    assert 'Amount spent (INR)' in meta_df.columns
    
    # Verify mapping works
    aliases = load_aliases()
    from src.inspection import map_columns
    
    leads_df = map_columns(leads_df, aliases)
    sales_df = map_columns(sales_df, aliases)
    meta_df = map_columns(meta_df, aliases)
    
    # Check leads mappings
    assert 'registration_date' in leads_df.columns
    assert 'campaign' in leads_df.columns
    assert 'ad_set' in leads_df.columns
    assert 'ad_creative' in leads_df.columns
    
    # Check meta mappings
    assert 'campaign' in meta_df.columns
    assert 'ad_set' in meta_df.columns
    assert 'ad' in meta_df.columns
    assert 'spend' in meta_df.columns
    assert 'Day' in meta_df.columns
    
    # Verify fallback price works in sales processing (order_amount is missing)
    from src.sales import process_sales
    
    settings = {
        'fallback_price': 123.45,
        'sale_date_source': 'Lead Registration Date',
        'payment_status_source': 'Treat All Imported Sales as Successful',
        'amount_source': 'Fallback Price Per Sale',
        'lead_start_date': '2026-01-01', 'ad_start_date': '2026-01-01',
        'lead_end_date': '2026-12-31', 'ad_end_date': '2026-12-31',
        'cutoff_date': '2026-01-01'
    }
    
    # We test the pipeline's date resolution logic as well
    metrics, ver_df, out = run_pipeline(leads_df, sales_df, [meta_df], settings, str(tmp_path / 'out.xlsx'))
    
    # The pipeline should have resolved sale_date from the lead (Jan 1, 2026)
    # The order_amount should be fallback (123.45)
    # The payment_status should be assumed_successful
    assert metrics['total_sales'] == 1
    assert metrics['total_revenue'] == 123.45
    
    # Read the output excel to check the columns
    all_sales = pd.read_excel(out, sheet_name="2. 📋 All Sales (Attributed)")
    assert 'sale_date_source' in all_sales.columns
    assert 'amount_source' in all_sales.columns
    assert 'payment_status_source' in all_sales.columns
    assert 'data_quality_warning' in all_sales.columns
    
    assert all_sales.iloc[0]['sale_date_source'] == 'lead_registration_date'
    assert 'fallback_price' in all_sales.iloc[0]['amount_source']
    assert all_sales.iloc[0]['payment_status_source'] == 'assumed_successful'
    assert 'Derived' in all_sales.iloc[0]['data_quality_warning']
