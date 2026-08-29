import pytest
import pandas as pd
import os
import openpyxl
from src.pipeline import run_pipeline

def test_excel_column_alignment(tmp_path):
    # Simulate a highly jagged sales dataframe with missing columns and different schema orders
    
    # 1. Leads
    leads_df = pd.DataFrame({
        'email': ['test1@example.com', 'test2@example.com'],
        'registration_date': ['2026-08-15', '2026-08-16'],
        'campaign': ['camp_a', 'camp_b'],
        'name': ['John', None],
        'phone': ['1234567890', None]
    })
    
    # 2. Sales (Missing some optional fields to test shifting)
    sales_df = pd.DataFrame({
        'email': ['test1@example.com', 'test2@example.com'],
        'sale_date': ['2026-08-17', '2026-08-18'],
        'order_amount': [100.0, 200.0],
        'name': ['John Doe', 'Jane Doe'],
        # intentionally skipping payment_status to test shifting
    })
    
    # 3. Meta
    meta_dfs = [pd.DataFrame({
        'campaign': ['camp_a', 'camp_b'],
        'spend': [50.0, 75.0],
        'Day': ['2026-08-15', '2026-08-15']
    })]
    
    settings = {
        'report_name': 'Test Report',
        'client_name': 'Test Client',
        'cutoff_date': '2026-01-01',
        'funnel_type': 'Paid',
        'fallback_price': 8999.0,
        'zero_roi_threshold': 5000.0,
        'currency': 'INR',
        'sale_date_source': 'Actual Sale Date',
        'payment_status_source': 'Treat All Imported Sales as Successful',
        'amount_source': 'Actual Order Amount',
        'custom_sale_date': None,
        'lead_start_date': '2026-08-01',
        'lead_end_date': '2026-08-31',
        'ad_start_date': '2026-08-01',
        'ad_end_date': '2026-08-31',
    }
    
    output_path = str(tmp_path / "test_report.xlsx")
    
    metrics, ver_df, xl_path = run_pipeline(leads_df, sales_df, meta_dfs, settings, output_path)
    
    wb = openpyxl.load_workbook(xl_path)
    
    # Check that unwanted sheets are removed
    assert "10. ❓ Unattributed Sales" not in wb.sheetnames
    assert "8. 💰 Free vs Paid Funnel" not in wb.sheetnames
    
    # Verify Campaign Summary
    ws_camp = wb["3. 📢 Campaign Summary"]
    headers = [cell.value for cell in ws_camp[1]]
    
    expected_headers = ['Campaign Name', 'Ad Account', 'Total Leads', 'Total Sales', 'Attributed Sales', 'Spend / Meta Spend', 'CPL', 'CAC', 'Revenue', 'Profit', 'ROAS', 'ROI', 'Conversion Rate', 'Price Per Sale', 'Funnel Type']
    
    assert headers == expected_headers, f"Headers did not match exactly: {headers}"
    
    # Check that the data values under 'Campaign Name' are indeed campaign names
    camp_idx = headers.index('Campaign Name')
    val1 = ws_camp.cell(row=2, column=camp_idx+1).value
    assert 'camp' in str(val1).lower(), f"Expected campaign name, got {val1}"
    
    # Check that 'Price Per Sale' has the expected 150 value (300 / 2)
    price_idx = headers.index('Price Per Sale')
    val_price = ws_camp.cell(row=2, column=price_idx+1).value
    assert val_price == 150, f"Expected 150 under Price Per Sale, got {val_price}"
    
    # Check All Sales sheet to ensure 'name' and 'phone' map correctly
    ws_sales = wb["2. 📋 All Sales (Attributed)"]
    s_headers = [cell.value for cell in ws_sales[1]]
    
    # Because sales_df was given 'name' but NOT 'phone', if they shift, the column for 'phone' will contain bad data.
    # We didn't provide 'phone' in sales_df, so it shouldn't exist or should be empty if the pipeline adds it.
    
    assert 'email' in s_headers
    assert 'name' in s_headers
    
    email_idx = s_headers.index('email') + 1
    name_idx = s_headers.index('name') + 1
    
    email_val = ws_sales.cell(row=2, column=email_idx).value
    name_val = ws_sales.cell(row=2, column=name_idx).value
    
    assert "@" in str(email_val)
    assert "Doe" in str(name_val)

    print("Excel mapping tests passed successfully!")
    
